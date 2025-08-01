from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Literal
import unicodedata
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from datetime import datetime


def get_suffix(file_name: str) -> str:
    path = Path(file_name)
    return path.suffix


def load_values(
    from_worksheet: Worksheet, 
    index: int,
    direction: Literal["row", "column"],
    start: int
) -> list[any]:
    worksheet = from_worksheet

    propagated_values = _propagate_merged_values(
        from_worksheet=worksheet,
        index=index + 1,
        direction=direction,
        start=start
    )

    return _remove_empty_values(propagated_values)


def _propagate_merged_values(
    from_worksheet: Worksheet, 
    index: int, 
    direction: Literal["row", "column"], 
    start: int
) -> list[str | None]:
    worksheet = from_worksheet
    merged_ranges = worksheet.merged_cells.ranges
    end = worksheet.max_column + 1 if direction == "row" else worksheet.max_row + 1
    values = []

    for row_or_column_index in range(start, end):
        cell = worksheet.cell(
            row=index if direction == "row" else row_or_column_index, 
            column=row_or_column_index if direction == "row" else index
        )

        value = cell.value

        if value is not None:
            values.append(value)
        else:
            for merged_range in merged_ranges:
                if cell.coordinate in merged_range:
                    top_left = worksheet.cell(
                        row=merged_range.min_row,
                        column=merged_range.min_col
                    )

                    values.append(top_left.value)
                    break
            else:
                values.append(None)

    return values


def _remove_empty_values(values: list[any]) -> list[any]:
    for i in range(len(values) - 1, -1, -1):
        if values[i] is not None:
            return values[:i + 1]
        
    return []


def normalize(text: str) -> str:
    if not isinstance(text, str):
        return ""
    
    text.strip().upper()

    return ''.join(
        c for c in unicodedata.normalize('NFD', text)
        if unicodedata.category(c) != 'Mn'
    )


def get_rows_and_columns_from(value: any, worksheet: Worksheet) -> list[tuple[int, int]]:
    coordinates = []

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value == value:
                coordinates.append(cell.coordinate)

    coordinates = sorted(coordinates, key=_sorting_key)

    rows_and_columns = []

    for coordinate in coordinates:
        column_letter, row_index = coordinate_from_string(coordinate)
        column_index = column_index_from_string(column_letter)
        rows_and_columns.append((row_index, column_index))

    return rows_and_columns


def _sorting_key(coordinate: str):
    letter = ''.join(filter(str.isalpha, coordinate))
    number = int(''.join(filter(str.isdigit, coordinate)))
    return (letter, number)


def join_sheets_vertically(worksheets: list[Worksheet]) -> Worksheet:
    output_workbook = Workbook()
    new_worksheet = output_workbook.active

    current_row = 1

    for i, worksheet in enumerate(worksheets):
        for j, row in enumerate(worksheet.iter_rows(values_only=True)):
            if i > 0 and j == 0:
                continue

            for column_index, value in enumerate(row, start=1):
                new_worksheet.cell(row=current_row, column=column_index, value=value)

            current_row += 1

    return new_worksheet


def get_date_from(text: str):
    for fmt in ("%d/%m/%Y", "%d:%m:%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    print("Wrong date format. Use dd/mm/yyyy, dd:mm:yyy or yyyy-mm-dd.")
    return None