from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import os
from typing import Literal
from .utils import normalize


def _load_distributors_sheet() -> Worksheet:
    file_path = os.path.join(os.path.dirname(__file__), "../../distribuidoras.xlsx")
    file_path = os.path.abspath(file_path)

    workbook = load_workbook(file_path, keep_links=False, read_only=True, data_only=True)
    return workbook.active


def get_column_info(unknown_column_name: str, known_column_name: str, known_value: str):
    worksheet = _load_distributors_sheet()

    header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    known_value_index = header.index(known_column_name)
    unknown_value_index = header.index(unknown_column_name)

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        known_cell = row[known_value_index]
        unknown_cell = row[unknown_value_index]

        if known_cell == known_value:
            return unknown_cell
        
    return None


def _load_acronyms(agent: Literal["Concessionária", "Permissionária"]) -> list[str]:
    distributors_sheet = _load_distributors_sheet()

    header = [cell.value for cell in next(distributors_sheet.iter_rows(min_row=1, max_row=1))]
    acronym_index = header.index("SIGLA")
    agent_index = header.index("AGENTE")

    filtered_acronyms = []

    for row in distributors_sheet.iter_rows(min_row=2, values_only=True):
        agent_value = row[agent_index]
        acronym_value = row[acronym_index]

        if agent_value == agent:
            filtered_acronyms.append(acronym_value)

    return filtered_acronyms


def _create_folders(agent: Literal["Concessionária", "Permissionária"]):
    base_path = os.path.join(os.path.dirname(__file__), "../../")
    base_path = os.path.abspath(base_path)

    distributors_path = os.path.join(base_path, f"{agent}s")
    distributors = _load_acronyms(agent=agent)

    for distributor in distributors:
        distributor_path = os.path.join(distributors_path, distributor)
        os.makedirs(distributor_path, exist_ok=True)

        for tariff_process in ["Ajuste EER ANGRA III", "Liminar abrace", "Reajuste", "Revisão", "Revisão Extraordinária", "Tarifas Iniciais"]:
            tariff_path = os.path.join(distributor_path, tariff_process)
            os.makedirs(tariff_path)
    

def create_all_folders():
    _create_folders("Concessionária")
    _create_folders("Permissionária")


def _load_value(column_name: str, acronym: str, from_sheet: Worksheet) -> any:
    worksheet = from_sheet
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header = list(header_row)

    try:
        aimed_column = header.index(column_name)
        acronym_column = header.index("SIGLA")
    except ValueError:
        raise ValueError(f"Column '{column_name}' or 'SIGLA' not found in header")
    
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        column_value = row[aimed_column]
        acronym_value = row[acronym_column]

        if normalize(acronym_value) ==  normalize(acronym):
            return column_value.strip() if isinstance(column_value, str) else column_value
    
    print(f"Sigla '{acronym}' não encontrada.")
    return None


def get_distributor_info(acronym: str) -> dict[str, any]:
    distributors_sheet = _load_distributors_sheet()

    name = _load_value(
        column_name='NOME',
        acronym=acronym,
        from_sheet=distributors_sheet
    )

    agent = _load_value(
        column_name='AGENTE',
        acronym=acronym,
        from_sheet=distributors_sheet
    )

    company_code = _load_value(
        column_name='CÓDIGO',
        acronym=acronym,
        from_sheet=distributors_sheet
    )

    agent_id = _load_value(
        column_name='ID AGENTE',
        acronym=acronym,
        from_sheet=distributors_sheet
    )

    concession_id = _load_value(
        column_name='ID CONCESSÃO',
        acronym=acronym,
        from_sheet=distributors_sheet
    )

    return {
        'name': name,
        'agent': agent,
        'company_code': company_code,
        'agent_id': agent_id,
        'concession_id': concession_id
    }