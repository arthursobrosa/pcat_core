from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Optional
from ..utils import load_values


def load_costs_sheet(workbook: Workbook) -> Optional[Worksheet]:
    tab_name = "CUSTOS"

    if tab_name not in workbook.sheetnames:
        return None

    costs_tab = workbook[tab_name]
    header_row = next(costs_tab.iter_rows(min_row=1, max_row=1, values_only=True))
    header = list(header_row)

    cost_info = load_values(
        from_worksheet=costs_tab,
        index=header.index("CUSTO"),
        direction="column",
        start=2
    )

    totals_indexes = _get_totals_indexes(cost_info)

    _remove_values_at(
        totals_indexes=totals_indexes,
        values=cost_info
    )

    length = len(cost_info)
    cost_info = _appended_values(cost_info)

    tariff_type_info = load_values(
        from_worksheet=costs_tab,
        index=header.index("TIPO TARIFA"),
        direction="column",
        start=2
    )

    _remove_values_at(
        totals_indexes=totals_indexes,
        values=tariff_type_info
    )

    tariff_type_info = _appended_values(tariff_type_info)

    cost_group_info = load_values(
        from_worksheet=costs_tab,
        index=header.index("GRUPO DE CUSTO"),
        direction="column",
        start=2
    )

    _remove_values_at(
        totals_indexes=totals_indexes,
        values=cost_group_info
    )

    cost_group_info = _appended_values(cost_group_info)

    cost_type_info = _load_cost_type_info(length)

    cost_type_values = _get_cost_type_values(
        from_worksheet=costs_tab,
        header=header,
        totals_indexes=totals_indexes
    )

    output_workbook = Workbook()
    new_worksheet = output_workbook.active

    new_worksheet.append([
        "TIPO TARIFA",
        "GRUPO DE CUSTO",
        "CUSTO",
        "TIPO DE CUSTO",
        "VALORES"
    ])

    for row in zip(tariff_type_info, cost_group_info, cost_info, cost_type_info, cost_type_values):
        new_worksheet.append(row)

    return new_worksheet


def _get_totals_indexes(values: list[any]) -> list[int]:
    totals = ["SUBTOTAL", "TOTAL", "TOTAL ABAS", "AVALIAÇÃO"]
    indexes = []

    for index in range(len(values)):
        value = values[index]

        if value in totals:
            indexes.append(index)

    return indexes


def _remove_values_at(totals_indexes: list[int], values: list[any]):
    for i, total_index in enumerate(totals_indexes):
        if i == 0:
            del values[total_index]
        else:
            del values[total_index - i]


def _appended_values(values: list[any], times: int = 3) -> list[any]:
    new_list = []

    for _ in range(times):
        new_list += values

    return new_list


def _load_cost_type_info(length: int) -> list[str]:
    economic_base = []
    financial_base = []
    cva = []

    for _ in range(length):
        economic_base.append("BASE ECONÔMICA")
        financial_base.append("BASE FINANCEIRA")
        cva.append("CVA")

    return economic_base + financial_base + cva


def _get_cost_type_values(from_worksheet: Worksheet, header: list[str], totals_indexes: list[int]):
    worksheet = from_worksheet

    economic_base_values = load_values(
        from_worksheet=worksheet,
        index=header.index("BASE ECONÔMICA"),
        direction="column",
        start=2
    )

    _remove_values_at(
        totals_indexes=totals_indexes,
        values=economic_base_values
    )

    financial_base_values = load_values(
        from_worksheet=worksheet,
        index=header.index("BASE FINANCEIRA"),
        direction="column",
        start=2
    )

    _remove_values_at(
        totals_indexes=totals_indexes,
        values=financial_base_values
    )

    cva_values = load_values(
        from_worksheet=worksheet,
        index=header.index("CVA"),
        direction="column",
        start=2
    )

    _remove_values_at(
        totals_indexes=totals_indexes,
        values=cva_values
    )

    return economic_base_values + financial_base_values + cva_values