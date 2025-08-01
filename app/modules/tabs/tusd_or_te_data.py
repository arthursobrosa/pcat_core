from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from enum import Enum
from typing import Optional
from itertools import groupby
from ..utils import load_values, get_rows_and_columns_from, join_sheets_vertically


class TusdOrTe(Enum):
    TUSD = 1
    TE = 2

    @property
    def main_tab(self) -> str:
        match self:
            case TusdOrTe.TUSD:
                return "TUSD"
            case TusdOrTe.TE:
                return "TE"
            
    @property
    def reference_tab(self) -> str:
        match self:
            case TusdOrTe.TUSD:
                return "TR TUSD"
            case TusdOrTe.TE:
                return "TR TE"
            
    @property
    def tariff_types(self) -> list[str]:
        match self:
            case TusdOrTe.TUSD:
                return ["TR TUSD", "TUSD BE", "TUSD BF", "TUSD CVA"]
            case TusdOrTe.TE:
                return ["TR TE", "TE BE", "TE BF", "TE CVA"]


def load_tusd_or_te_sheet(workbook: Workbook, tusd_or_te: TusdOrTe) -> Optional[Worksheet]:
    tab_name = tusd_or_te.main_tab

    if tab_name not in workbook.sheetnames:
        return None

    tab = workbook[tab_name]
    header_row = next(tab.iter_rows(min_row=1, max_row=1, values_only=True))
    header = list(header_row)

    length = _get_length(
        workbook=workbook,
        header=header,
        reference_tab=tusd_or_te.reference_tab
    )

    main_sheet = _load_main_sheet(
        workbook=workbook,
        header=header,
        length=length,
        tusd_or_te=tusd_or_te
    )

    remaining_sheets = _get_remaining_sheets(
        workbook=workbook,
        length=length,
        tusd_or_te=tusd_or_te
    )

    output_workbook = Workbook()
    new_worksheet = output_workbook.active

    for row in main_sheet.iter_rows(values_only=True):
        new_worksheet.append(list(row))

    num_cols_main_sheet = main_sheet.max_column

    remaining_sheets_data = []

    for sheet in remaining_sheets:
        for row in sheet.iter_rows(values_only=True):
            remaining_sheets_data.append(list(row))

    for i, row_data in enumerate(remaining_sheets_data, start=1):
        for j, value in enumerate(row_data, start=1):
            new_worksheet.cell(row=i, column=num_cols_main_sheet + j, value=value)

    return new_worksheet


def _load_main_sheet(
    workbook: Workbook, 
    header: list[str], 
    length: int, 
    tusd_or_te: TusdOrTe
) -> Worksheet:
    tariff_type_info = _load_tariff_type_info(
        length=length,
        tariff_types=tusd_or_te.tariff_types
    )

    subgroup_info = _load_info_at(
        workbook=workbook,
        header=header,
        tariff_types=tusd_or_te.tariff_types,
        column_name="SUBGRUPO",
        insert_new_row=True
    )

    modality_info = _load_info_at(
        workbook=workbook,
        header=header,
        tariff_types=tusd_or_te.tariff_types,
        column_name="MODALIDADE",
        insert_new_row=True
    )

    class_info = _load_info_at(
        workbook=workbook,
        header=header,
        tariff_types=tusd_or_te.tariff_types,
        column_name="CLASSE",
        insert_new_row=True
    )

    subclass_info = _load_info_at(
        workbook=workbook,
        header=header,
        tariff_types=tusd_or_te.tariff_types,
        column_name="SUBCLASSE",
        insert_new_row=True
    )

    detail_info = _load_info_at(
        workbook=workbook,
        header=header,
        tariff_types=tusd_or_te.tariff_types,
        column_name="DETALHE",
        insert_new_row=True
    )

    uc_column_name = workbook[tusd_or_te.main_tab]["F1"].value

    uc_info = _load_info_at(
        workbook=workbook,
        header=header,
        tariff_types=tusd_or_te.tariff_types,
        column_name=uc_column_name,
        insert_new_row=True
    )

    post_info = _load_info_at(
        workbook=workbook,
        header=header,
        tariff_types=tusd_or_te.tariff_types,
        column_name="POSTO",
        insert_new_row=True
    )

    unity_info = _load_info_at(
        workbook=workbook,
        header=header,
        tariff_types=tusd_or_te.tariff_types,
        column_name="UNIDADE",
        insert_new_row=True
    )

    output_workbook = Workbook()
    new_worksheet = output_workbook.active

    new_worksheet.append([
        "TIPO DE TARIFA",
        "SUBGRUPO",
        "MODALIDADE",
        "CLASSE",
        "SUBCLASSE",
        "DETALHE",
        uc_column_name,
        "POSTO",
        "UNIDADE"
    ])

    for row in zip(tariff_type_info, subgroup_info, modality_info, class_info, subclass_info, detail_info, uc_info, post_info, unity_info):
        new_worksheet.append(row)

    return new_worksheet
    

def _load_tariff_type_info(length: int, tariff_types: list[str]) -> list[str]:
    all_info = [""]

    for tariff_type in tariff_types:
        for _ in range(length):
            all_info.append(tariff_type)

    return all_info


def _get_length(workbook: Workbook, header: list[str], reference_tab: str) -> int:
    values = load_values(
        from_worksheet=workbook[reference_tab],
        index=header.index("SUBGRUPO"),
        direction="column",
        start=5
    )

    return len(values)

    
def _load_info_at(
    workbook: Workbook, 
    header: list[str], 
    tariff_types: list[str], 
    column_name: str, 
    insert_new_row: bool = False
) -> list[any]:
    all_info = [""] if insert_new_row else []

    for tariff_type in tariff_types:
        worksheet = workbook[tariff_type]

        info = load_values(
            from_worksheet=worksheet,
            index=header.index(column_name),
            direction="column",
            start=5
        )

        all_info += info

    return all_info


def _get_remaining_sheets(workbook: Workbook, length: int, tusd_or_te: TusdOrTe) -> list[Worksheet]:
    sheets = []

    for index, tariff_type in enumerate(tusd_or_te.tariff_types):
        tab_sheet = _get_sheet_from_tab(
            workbook=workbook,
            reference_tab=tusd_or_te.reference_tab,
            tab_name=tariff_type,
            length=length,
            first_tab= index == 0
        )

        sheets.append(tab_sheet)

    return sheets


def _get_sheet_from_tab(
    workbook: Workbook, 
    reference_tab: str, 
    tab_name: str, 
    length: int, 
    first_tab: bool
) -> Worksheet:
    remaining_header = _get_remaining_header(
        workbook=workbook,
        reference_tab=reference_tab
    )

    output_workbook = Workbook()
    new_worksheet = output_workbook.active

    for col_index, title in enumerate(remaining_header, start=1):
        new_worksheet.cell(row=1, column=col_index, value=title)

    worksheet = workbook[tab_name]
    header_row = next(worksheet.iter_rows(min_row=3, max_row=3, values_only=True))
    header = list(header_row)
    column_index = header.index(remaining_header[0])
    all_values = []

    for column_name in remaining_header:
        if isinstance(column_name, str):
            values = load_values(
                from_worksheet=worksheet,
                index=column_index,
                direction="column",
                start=4
            )

            values = values[:length+1]
            all_values.append(values)

        column_index += 1

    for row_index, row_values in enumerate(zip(*all_values), start=2):
        for col_index, value in enumerate(row_values, start=1):
            new_worksheet.cell(row=row_index, column=col_index, value=value)

    if not first_tab:
        data = list(new_worksheet.iter_rows(min_row=3, values_only=True))
        new_worksheet.delete_rows(1, new_worksheet.max_row)

        for row_index, row in enumerate(data, start=1):
            for col_index, value in enumerate(row, start=1):
                new_worksheet.cell(row=row_index, column=col_index, value=value)

    return new_worksheet
    

def _get_remaining_header(workbook: Workbook, reference_tab: str) -> list[any]:
    worksheet = workbook[reference_tab]

    headers = load_values(
        from_worksheet=worksheet,
        index=2,
        direction="row",
        start=12
    )

    return _filtered_header(headers)


def _filtered_header(values: list[any]) -> list[any]:
    filtered_values = []

    for value in values:
        if value is None:
            break

        filtered_values.append(value)

    return filtered_values


def create_mixed_tusd_or_te_worksheet(
    workbooks: list[Workbook], 
    tusd_or_te: TusdOrTe,
    output_workbook: Workbook
):
    headers = []
    worksheets = []

    for workbook in workbooks:
        if tusd_or_te.main_tab not in workbook.sheetnames:
            continue

        file_worksheet = workbook[tusd_or_te.main_tab]
        worksheets.append(file_worksheet)

        header_row = next(file_worksheet.iter_rows(min_row=1, max_row=1, min_col=18, values_only=True))
        header = list(header_row)
        headers += header

    files_dict = _get_files_dict(
        worksheets=worksheets,
        headers=set(headers)
    )

    headers_and_tariffs = _get_headers_and_tariffs(
        worksheets=worksheets,
        files_dict=files_dict
    )

    files_columns = _get_files_columns(
        worksheets=worksheets,
        headers_and_tariffs=headers_and_tariffs
    )

    data = []

    for i in range(len(headers_and_tariffs)):
        column_values = []

        for j, worksheet in enumerate(worksheets):
            column = files_columns[worksheet][i]

            if column is None:
                if j == 0:
                    column_values.append(headers_and_tariffs[i][0])
                    column_values.append(headers_and_tariffs[i][1])

                for row_index in range(3, worksheet.max_row + 1):
                    column_values.append("Nao se aplica")

                continue

            values = load_values(
                from_worksheet=worksheet,
                index=column - 1,
                direction="column",
                start=1 if j == 0 else 3
            )

            column_values += values

        data.append(column_values)

    cleaned_worksheet = _get_cleaned_worksheet(worksheets=worksheets)
    data_worksheet = _get_worksheet_from_data(data)
    final_worksheet = _create_side_by_side_worksheet(worksheets=[cleaned_worksheet, data_worksheet])

    new_worksheet = output_workbook.create_sheet(title=tusd_or_te.main_tab)

    for row in final_worksheet.iter_rows(values_only=True):
        new_worksheet.append(row)


def _get_files_dict(worksheets: list[Worksheet], headers: set[str]) -> dict[Worksheet, list[tuple[int, int]]]:
    files_dict = {}

    for worksheet in worksheets:
        all_rows_and_columns = []

        for header in headers:
            rows_and_columns = get_rows_and_columns_from(
                value=header,
                worksheet=worksheet
            )

            rows_and_columns = [rc for rc in rows_and_columns if rc[0] == 1]
            all_rows_and_columns += rows_and_columns

        all_rows_and_columns = sorted(all_rows_and_columns, key=lambda x: x[1])
        files_dict[worksheet] = all_rows_and_columns

    return files_dict


def _get_headers_and_tariffs(worksheets: list[Worksheet], files_dict: dict[Worksheet, list[tuple[int, int]]]) -> list[tuple[any, any]]:
    headers_and_tariffs = []

    for worksheet in worksheets:
        rows_and_columns = files_dict[worksheet]

        for row_and_column in rows_and_columns:
            values = load_values(
                from_worksheet=worksheet,
                index=row_and_column[1] - 1,
                direction="column",
                start=row_and_column[0]
            )

            header_and_tariff = (values[0], values[1])
            headers_and_tariffs.append(header_and_tariff)

    headers_and_tariffs = sorted(headers_and_tariffs, key=lambda x: x[0])
    unrepeated_values = []

    for _, group in groupby(headers_and_tariffs, key=lambda x: x[0]):
        exclusive_group = set(group)
        unrepeated_values.extend(sorted(exclusive_group, key=lambda x: (x[1] == 'SUBTOTAL', x)))

    headers_and_tariffs = unrepeated_values
    return headers_and_tariffs


def _get_files_columns(worksheets: list[Worksheet], headers_and_tariffs: list[tuple[any, any]]) -> dict[Worksheet, list[int]]:
    files_columns = {}

    for worksheet in worksheets:
        all_common_columns = []

        for header_and_tariff in headers_and_tariffs:
            header_rows_and_columns = get_rows_and_columns_from(
                value=header_and_tariff[0],
                worksheet=worksheet
            )

            tariff_rows_and_columns = get_rows_and_columns_from(
                value=header_and_tariff[1],
                worksheet=worksheet
            )

            header_columns = [rc[1] for rc in header_rows_and_columns if rc[0] == 1]
            tariff_columns = [rc[1] for rc in tariff_rows_and_columns if rc[0] == 2]
            common_columns = list(set(header_columns) & set(tariff_columns))

            if len(common_columns) == 0:
                all_common_columns.append(None)
                continue

            all_common_columns += common_columns

        files_columns[worksheet] = all_common_columns

    return files_columns


def _get_cleaned_worksheet(worksheets: list[Worksheet]) -> Worksheet:
    for i, worksheet in enumerate(worksheets):
        worksheet.delete_cols(idx=18, amount=worksheet.max_column - 17)

        if i > 0:
            worksheet.delete_rows(idx=1)

    return join_sheets_vertically(worksheets=worksheets)


def _get_worksheet_from_data(data: list[list[any]]) -> Worksheet:
    output_workbook = Workbook()
    new_worksheet = output_workbook.active

    for column_index, column in enumerate(data, start=1):
        for row_index, value in enumerate(column, start=1):
            new_worksheet.cell(row=row_index, column=column_index, value=value)

    return new_worksheet


def _create_side_by_side_worksheet(worksheets: list[Worksheet]) -> Worksheet:
    output_workbook = Workbook()
    new_worksheet = output_workbook.active

    current_column = 1

    for worksheet in worksheets:
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        for i in range(1, max_row + 1):
            for j in range(1, max_col + 1):
                value = worksheet.cell(row=i, column=j).value
                new_worksheet.cell(row=i, column=current_column + j - 1, value=value)

        current_column += max_col

    return new_worksheet