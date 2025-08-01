from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Literal, Optional
from ..utils import load_values


def load_tusd_or_te_market_sheet(workbook: Workbook, tusd_or_te: Literal["TUSD", "TE"]) -> Optional[Worksheet]:
    tab_name = f"MERCADO {tusd_or_te}"

    if tab_name not in workbook.sheetnames:
        return None

    tab = workbook[tab_name]
    header_row = next(tab.iter_rows(min_row=1, max_row=1, values_only=True))
    header = list(header_row)

    subgroup = load_values(
        from_worksheet=tab,
        index=header.index("SUBGRUPO"),
        direction="column",
        start=2
    )

    modality = load_values(
        from_worksheet=tab,
        index=header.index("MODALIDADE"),
        direction="column",
        start=2
    )

    class_values = load_values(
        from_worksheet=tab,
        index=header.index("CLASSE"),
        direction="column",
        start=2
    )

    subclass = load_values(
        from_worksheet=tab,
        index=header.index("SUBCLASSE"),
        direction="column",
        start=2
    )

    detail = load_values(
        from_worksheet=tab,
        index=header.index("DETALHE"),
        direction="column",
        start=2
    )

    consumer_unit = load_values(
        from_worksheet=tab,
        index=header.index("NOME UC"),
        direction="column",
        start=2
    )

    post = load_values(
        from_worksheet=tab,
        index=header.index("POSTO"),
        direction="column",
        start=2
    )

    unity = load_values(
        from_worksheet=tab,
        index=header.index("UNIDADE"),
        direction="column",
        start=2
    )

    reference_market = load_values(
        from_worksheet=tab,
        index=header.index("SOMA MERCADO"),
        direction="column",
        start=2
    )

    output_workbook = Workbook()
    new_worksheet = output_workbook.active

    new_worksheet.append([
        "SUBGRUPO",
        "MODALIDADE",
        "CLASSE",
        "SUBCLASSE",
        "DETALHE",
        "UC",
        "POSTO",
        "UNIDADE",
        "MERCADO DE REFERÊNCIA"
    ])

    for row in zip(subgroup, modality, class_values, subclass, detail, consumer_unit, post, unity, reference_market):
        new_worksheet.append(row)

    last_row = new_worksheet.max_row
    new_worksheet.delete_rows(last_row)

    return new_worksheet