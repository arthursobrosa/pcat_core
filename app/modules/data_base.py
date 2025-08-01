from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook, Workbook
from typing import Literal, Optional
from datetime import datetime
import os
from tqdm import tqdm
from .distributor_info import get_distributor_info
from .tabs.costs_data import load_costs_sheet
from .tabs.tusd_or_te_market_data import load_tusd_or_te_market_sheet
from .tabs.tusd_or_te_data import load_tusd_or_te_sheet, TusdOrTe, create_mixed_tusd_or_te_worksheet
from .tabs.effect_data import load_effect_sheet
from .tabs.reh_tables_data import load_reh_tables_sheet
from .utils import get_date_from, get_suffix


def merge_last_dbs():
    base_path = os.path.join(os.path.dirname(__file__), "../../")
    base_path = os.path.abspath(base_path)

    data_base_path = os.path.join(base_path, "Banco de Dados")

    file_names = [
        name for name in os.listdir(data_base_path)
        if (name.endswith(".xlsx") or name.endswith(".xlsm")) 
        and not name.startswith("~$")
    ]

    file_paths = []

    for file_name in file_names:
        file_path = os.path.join(data_base_path, file_name)
        file_paths.append(file_path)

    output_name = os.path.join(data_base_path, "BANCO_Geral.xlsx")

    _mix_db_files(
        file_paths=file_paths,
        output_name=output_name
    )


def process_data_base(agent: Literal["Concessionária", "Permissionária"]):
    base_path = os.path.join(os.path.dirname(__file__), "../../")
    base_path = os.path.abspath(base_path)

    distributors_path = os.path.join(base_path, f"{agent}s")

    distributors = [
        name for name in os.listdir(distributors_path)
        if os.path.isdir(os.path.join(distributors_path, name))
    ]

    distributors.sort()

    all_file_paths = []

    for distributor in tqdm(distributors, desc="Processando distribuidoras..."):
        distributor_path = os.path.join(distributors_path, distributor)

        data_base_path = os.path.join(distributor_path, "Banco de Dados")

        if not os.path.isdir(data_base_path):
            continue

        file_names = [
            name for name in os.listdir(data_base_path)
            if (name.endswith(".xlsx") or name.endswith(".xlsm")) 
            and not name.startswith("~$")
        ]

        if not file_names or len(file_names) != 1:
            continue

        file_path = os.path.join(data_base_path, file_names[0])
        all_file_paths.append(file_path)

    if all_file_paths:
        output_folder_path = os.path.join(base_path, "Banco de Dados")
        os.makedirs(output_folder_path, exist_ok=True)

        output_path = os.path.join(output_folder_path, f"BANCO_{agent}s.xlsx")

        _mix_db_files(
            file_paths=all_file_paths,
            output_name=output_path
        )


def process_workbooks(agent: Literal["Concessionária", "Permissionária"]):
    base_path = os.path.join(os.path.dirname(__file__), "../../")
    base_path = os.path.abspath(base_path)

    distributors_path = os.path.join(base_path, f"{agent}s")

    distributors = [
        name for name in os.listdir(distributors_path)
        if os.path.isdir(os.path.join(distributors_path, name))
    ]

    distributors.sort()

    for distributor in tqdm(distributors, desc="Processando distribuidoras..."):
        distributor_path = os.path.join(distributors_path, distributor)

        temp_file_paths = []

        for type in ["Ajuste EER ANGRA III", "Liminar abrace", "Reajuste", "Revisão", "Revisão Extraordinária", "Tarifas Iniciais"]:
            type_path = os.path.join(distributor_path, type)

            file_names = [
                name for name in os.listdir(type_path)
                if (name.endswith(".xlsx") or name.endswith(".xlsm")) 
                and not name.startswith("~$")
            ]

            for file_name in tqdm(file_names, desc=f"{distributor} - {type}", leave=False):
                file_path = os.path.join(type_path, file_name)
                file_workbook = load_workbook(file_path, data_only=True)

                suffix = get_suffix(file_name)
                file_name_without_suffix = file_name.replace(suffix, "")
                parts = file_name_without_suffix.split("_")
                process_date_str = parts[len(parts) - 1]
                process_date = get_date_from(process_date_str)

                try:
                    new_workbook = _filtered_workbook(
                        workbook=file_workbook,
                        acronym=distributor,
                        tariff_process=type,
                        process_date=process_date
                    )

                    temp_path = file_path.replace(suffix, f"_temp{suffix}")
                    new_workbook.save(temp_path)
                    temp_file_paths.append(temp_path)
                except Exception as error:
                    print(f"\nFalha ao filtrar planilha em {file_path}: {str(error)}") 

        if temp_file_paths:
            output_folder_path = os.path.join(distributor_path, "Banco de Dados")
            os.makedirs(output_folder_path, exist_ok=True)

            output_path = os.path.join(output_folder_path, f"{distributor}_BANCO.xlsx")

            _mix_db_files(
                file_paths=temp_file_paths,
                output_name=output_path
            )

            for temp_file in temp_file_paths:
                os.remove(temp_file)


def _filtered_workbook(
    workbook: Workbook,
    acronym: str, 
    tariff_process: Literal["Ajuste EER ANGRA III", "Liminar abrace", "Reajuste", "Revisão", "Revisão Extraordinária", "Tarifas Iniciais"],
    process_date: any
) -> Workbook:
    distributor_info = get_distributor_info(acronym=acronym)
    distributor_info = {
        'Nome': distributor_info['name'],
        'Sigla': acronym,
        'Concessionária/Permissionária': distributor_info['agent'],
        'Código da Empresa': distributor_info['company_code'],
        'ID Agente': distributor_info['agent_id'],
        'ID Concessão': distributor_info['concession_id'],
        'Processo Tarifário': tariff_process,
        'Data do processo tarifário em processamento': process_date
    }

    distributor_header = list(distributor_info.keys())

    new_workbook = Workbook()
    default_sheet = new_workbook.active
    new_workbook.remove(default_sheet)

    _create_db_tab(
        distributor_info=distributor_info,
        distributor_header=distributor_header,
        workbook=new_workbook,
        worksheet=load_costs_sheet(workbook=workbook),
        tab_name='CUSTOS'
    )

    _create_db_tab(
        distributor_info=distributor_info,
        distributor_header=distributor_header,
        workbook=new_workbook,
        worksheet=load_tusd_or_te_market_sheet(workbook=workbook, tusd_or_te="TUSD"),
        tab_name='MERCADO TUSD'
    )

    _create_db_tab(
        distributor_info=distributor_info,
        distributor_header=distributor_header,
        workbook=new_workbook,
        worksheet=load_tusd_or_te_sheet(workbook=workbook, tusd_or_te=TusdOrTe.TUSD),
        tab_name='TUSD',
        hide_first_line=True
    )

    _create_db_tab(
        distributor_info=distributor_info,
        distributor_header=distributor_header,
        workbook=new_workbook,
        worksheet=load_tusd_or_te_market_sheet(workbook=workbook, tusd_or_te="TE"),
        tab_name='MERCADO TE'
    )

    _create_db_tab(
        distributor_info=distributor_info,
        distributor_header=distributor_header,
        workbook=new_workbook,
        worksheet=load_tusd_or_te_sheet(workbook=workbook, tusd_or_te=TusdOrTe.TE),
        tab_name='TE',
        hide_first_line=True
    )

    _create_db_tab(
        distributor_info=distributor_info,
        distributor_header=distributor_header,
        workbook=new_workbook,
        worksheet=load_effect_sheet(workbook=workbook),
        tab_name='EFEITO'
    )

    _create_db_tab(
        distributor_info=distributor_info,
        distributor_header=distributor_header,
        workbook=new_workbook,
        worksheet=load_reh_tables_sheet(workbook=workbook),
        tab_name="TABELAS REH"
    )

    if len(new_workbook.sheetnames) == 0:
        new_workbook.create_sheet(title="Sheet")

    return new_workbook


def _add_header_rows(header_rows: list[any], to_sheet: Worksheet):
    worksheet = to_sheet

    for row_offset, row in enumerate(header_rows, start=1):
        for column_offset, cell in enumerate(row, start=1):
            target_row = row_offset
            target_column = column_offset

            value = getattr(cell, "value", None)
            worksheet.cell(row=target_row, column=target_column).value = value


def _mix_db_files(
    file_paths: list[str], 
    output_name: str,
    header_max_row: int = 1
):
    if not file_paths:
        print(f"Lista de caminhos de arquivos vazia (iria para {output_name})")
        return
    
    file_workbooks = [
        load_workbook(file_path, keep_links=False, read_only=True, data_only=True)
        for file_path in file_paths
    ]

    max_row_per_sheet = 1048576

    output_workbook = Workbook()
    output_worksheet = output_workbook.active
    output_worksheet.title = "BANCO DE DADOS"

    current_sheet = output_worksheet
    current_row_count = header_max_row
    sheet_index = 0

    header_rows = [] 

    for file_index, file_workbook in enumerate(file_workbooks):
        file_worksheet = file_workbook.active

        if file_index == 0:
            header_rows = list(file_worksheet.iter_rows(min_row=1, max_row=header_max_row, values_only=False))
            _add_header_rows(header_rows, to_sheet=current_sheet)

        min_row = header_max_row + 1
        max_row = file_worksheet.max_row

        for row in file_worksheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
            if current_row_count >= max_row_per_sheet:
                sheet_index += 1
                new_sheet_title = f"BANCO DE DADOS - Ext {sheet_index}"
                current_sheet = output_workbook.create_sheet(title=new_sheet_title)
                _add_header_rows(header_rows, to_sheet=current_sheet)
                current_row_count = header_max_row

            current_sheet.append(row)
            current_row_count += 1

    output_workbook.save(output_name)


def _create_db_tab(
    distributor_info: dict[str, any], 
    distributor_header: list[str], 
    workbook: Workbook, 
    worksheet: Optional[Worksheet], 
    tab_name: str, 
    hide_first_line: bool = False
):
    if not worksheet:
        return

    new_worksheet = workbook.create_sheet(title=tab_name)
    worksheet_header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    is_worksheet_empty = all(cell is None or str(cell).strip() == "" for cell in worksheet_header)

    if is_worksheet_empty:
        return

    new_worksheet.append(distributor_header + worksheet_header)

    max_row = worksheet.max_row
    max_col = worksheet.max_column

    counter = 0

    for row in worksheet.iter_rows(min_row=2, max_row=max_row, max_col=max_col, values_only=True):
        if not all(cell is None for cell in row):
            if hide_first_line and counter == 0:
                length = len(distributor_info.values())
                empty_row = []

                for _ in range(length):
                    empty_row.append("")

                new_row = empty_row + list(row)
                new_worksheet.append(new_row)
                counter += 1

                continue

            new_row = list(distributor_info.values()) + list(row)
            new_worksheet.append(new_row)